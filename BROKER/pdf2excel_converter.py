#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BROKER - PDF to Excel Converter
28회 공통 문제 PDF에서 엑셀 QUESTION 필드로 변환하는 도구

Author: AI Assistant (Seo Daeri)
Date: 2024-12
Purpose: PDF 텍스트 추출 및 엑셀 자동 입력
"""

import os
import re
import pandas as pd
import PyPDF2
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from datetime import datetime
import logging

class PDF2ExcelConverter:
    def __init__(self):
        self.setup_logging()
        self.excel_file = "GEP_MASTER_DB_V1.0.xlsx"
        self.pdf_file = "28회(2022)_공통(보험관계법령 등).pdf"
        
    def setup_logging(self):
        """로깅 설정"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('pdf_converter.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def extract_text_from_pdf(self, pdf_path):
        """PDF에서 텍스트 추출"""
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
                return text
        except Exception as e:
            self.logger.error(f"PDF 텍스트 추출 실패: {e}")
            return None
    
    def parse_questions_from_text(self, text):
        """텍스트에서 문제 파싱 (줄바꿈 기반)"""
        questions = []
        
        # 문제 번호 패턴 (1., 2., 3. 등)
        question_pattern = r'(\d+)\.\s*(.*?)(?=\d+\.|$)'
        matches = re.findall(question_pattern, text, re.DOTALL)
        
        for match in matches:
            question_num = int(match[0])
            question_text = match[1].strip()
            
            # 줄바꿈 처리: 문제끝, ①앞, ②앞, ③앞, ④앞
            formatted_text = self.format_question_text(question_text)
            
            question_data = {
                'question_number': question_num,
                'formatted_text': formatted_text,
                'original_text': question_text
            }
            
            questions.append(question_data)
        
        return questions
    
    def format_question_text(self, text):
        """문제 텍스트 포맷팅 (줄바꿈 처리)"""
        # 문제 끝에서 줄바꿈 추가
        # ①, ②, ③, ④ 앞에서 줄바꿈 추가
        
        # 선택지 패턴 찾기
        option_patterns = [
            r'①', r'②', r'③', r'④'
        ]
        
        formatted_text = text
        
        # 각 선택지 앞에 줄바꿈 추가
        for pattern in option_patterns:
            formatted_text = re.sub(pattern, f'\n{pattern}', formatted_text)
        
        # 문제 끝 부분에서 줄바꿈 추가 (선택지 시작 전)
        # 문제 지문과 선택지 사이에 줄바꿈 추가
        if '①' in formatted_text:
            parts = formatted_text.split('①', 1)
            if len(parts) == 2:
                question_part = parts[0].strip()
                options_part = '①' + parts[1]
                formatted_text = f"{question_part}\n{options_part}"
        
        return formatted_text
    
    def load_excel_file(self):
        """엑셀 파일 로드"""
        try:
            workbook = load_workbook(self.excel_file)
            worksheet = workbook.active
            return workbook, worksheet
        except Exception as e:
            self.logger.error(f"엑셀 파일 로드 실패: {e}")
            return None, None
    
    def find_target_cells(self, worksheet, target_round=28, target_layer1="관계법령"):
        """28회 관계법령 문제의 대상 셀 찾기"""
        target_cells = []
        
        for row in range(2, worksheet.max_row + 1):
            eround = worksheet.cell(row=row, column=5).value  # E열: EROUND
            layer1 = worksheet.cell(row=row, column=6).value  # F열: LAYER1
            
            if eround == target_round and str(layer1) == target_layer1:
                qnum = worksheet.cell(row=row, column=9).value  # I열: QNUM
                question_cell = worksheet.cell(row=row, column=11)  # K열: QUESTION
                
                target_cells.append({
                    'row': row,
                    'qnum': qnum,
                    'question_cell': question_cell
                })
                
                # 디버깅: 처음 5개만 출력
                if len(target_cells) <= 5:
                    self.logger.info(f"대상 셀 발견: 행={row}, EROUND={eround}, LAYER1={layer1}, QNUM={qnum}")
        
        self.logger.info(f"총 {len(target_cells)}개 대상 셀 발견")
        return target_cells
    
    def update_excel_with_questions(self, questions, target_cells):
        """엑셀 파일에 문제 업데이트"""
        workbook, worksheet = self.load_excel_file()
        if not workbook:
            return False
        
        updated_count = 0
        errors = []
        
        for question in questions:
            qnum = question['question_number']
            
            # 해당 문제 번호의 셀 찾기
            target_cell = None
            for cell_info in target_cells:
                if str(cell_info['qnum']) == str(qnum):
                    target_cell = cell_info
                    break
            
            if target_cell:
                try:
                    # 디버깅: 업데이트 전 값 확인
                    if qnum == 1:
                        self.logger.info(f"문제 1 업데이트 전 값: {target_cell['question_cell'].value}")
                        self.logger.info(f"문제 1 새 값 길이: {len(question['formatted_text'])}")
                        self.logger.info(f"문제 1 새 값 미리보기: {question['formatted_text'][:200]}...")
                    
                    # QUESTION 셀에 포맷된 텍스트 입력
                    target_cell['question_cell'].value = question['formatted_text']
                    
                    # 디버깅: 업데이트 후 값 확인
                    if qnum == 1:
                        self.logger.info(f"문제 1 업데이트 후 값: {target_cell['question_cell'].value}")
                    
                    updated_count += 1
                    self.logger.info(f"문제 {qnum} 업데이트 완료")
                        
                except Exception as e:
                    error_msg = f"문제 {qnum} 업데이트 실패: {e}"
                    errors.append(error_msg)
                    self.logger.error(error_msg)
            else:
                error_msg = f"문제 {qnum}에 해당하는 셀을 찾을 수 없음"
                errors.append(error_msg)
                self.logger.warning(error_msg)
                # 디버깅: 대상 셀들의 QNUM 값들 출력
                if qnum <= 5:  # 처음 5개만
                    qnum_values = [cell_info['qnum'] for cell_info in target_cells[:10]]
                    self.logger.info(f"대상 셀 QNUM 값들 (처음 10개): {qnum_values}")
        
        # 엑셀 파일 저장
        try:
            workbook.save(self.excel_file)
            self.logger.info(f"엑셀 파일 저장 완료: {updated_count}개 문제 업데이트")
            
            # 저장 후 검증
            verification_df = pd.read_excel(self.excel_file)
            verification_rows = verification_df[(verification_df['EROUND'] == 28) & (verification_df['LAYER1'] == '관계법령')]
            non_empty_questions = verification_rows['QUESTION'].notna().sum()
            self.logger.info(f"저장 후 검증: {non_empty_questions}개 문제에 내용이 있음")
            
        except Exception as e:
            self.logger.error(f"엑셀 파일 저장 실패: {e}")
            return False
        
        return {
            'success': True,
            'updated_count': updated_count,
            'total_questions': len(questions),
            'errors': errors
        }
    
    def run_conversion(self):
        """메인 변환 프로세스"""
        self.logger.info("=== PDF2EXCEL 변환 시작 ===")
        
        # 1. PDF 텍스트 추출
        self.logger.info(f"PDF 파일 읽기: {self.pdf_file}")
        text = self.extract_text_from_pdf(self.pdf_file)
        if not text:
            return {'success': False, 'error': 'PDF 텍스트 추출 실패'}
        
        # 2. 문제 파싱
        self.logger.info("문제 파싱 시작")
        questions = self.parse_questions_from_text(text)
        self.logger.info(f"추출된 문제 수: {len(questions)}")
        
        # 3. 대상 셀 찾기
        workbook, worksheet = self.load_excel_file()
        if not workbook:
            return {'success': False, 'error': '엑셀 파일 로드 실패'}
        
        target_cells = self.find_target_cells(worksheet)
        self.logger.info(f"대상 셀 수: {len(target_cells)}")
        
        # 4. 엑셀 업데이트
        self.logger.info("엑셀 파일 업데이트 시작")
        result = self.update_excel_with_questions(questions, target_cells)
        
        return result
    
    def create_test_report(self, result):
        """테스트 결과 보고서 생성"""
        report = {
            'timestamp': datetime.now().isoformat(),
            'test_type': 'PDF2EXCEL 변환 테스트',
            'target_file': self.pdf_file,
            'result': result,
            'summary': {
                'success': result.get('success', False),
                'updated_count': result.get('updated_count', 0),
                'total_questions': result.get('total_questions', 0),
                'error_count': len(result.get('errors', []))
            }
        }
        
        # JSON 파일로 저장
        with open('test_result.json', 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        return report

def main():
    """메인 실행 함수"""
    converter = PDF2ExcelConverter()
    
    # 변환 실행
    result = converter.run_conversion()
    
    # 결과 출력
    if result.get('success'):
        print(f"✅ 변환 성공!")
        print(f"📊 업데이트된 문제: {result.get('updated_count')}개")
        print(f"📋 총 문제 수: {result.get('total_questions')}개")
        
        if result.get('errors'):
            print(f"⚠️ 오류 수: {len(result.get('errors'))}개")
            for error in result.get('errors')[:5]:  # 처음 5개만 출력
                print(f"  - {error}")
    else:
        print(f"❌ 변환 실패: {result.get('error')}")
    
    # 테스트 보고서 생성
    report = converter.create_test_report(result)
    print(f"📄 테스트 보고서 생성 완료: test_result.json")

if __name__ == "__main__":
    main()
